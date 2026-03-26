from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from .base import detect_languages, make_document, normalize_text
from ..schemas import BlockRecord, CitationAnchor, NormalizedSegment


class XlsxParser:
    parser_name = "xlsx_parser"

    def parse(self, path: Path, relative_path: str):
        workbook = load_workbook(path, data_only=True)
        document = make_document(path, relative_path, path.suffix.lower().lstrip("."), parser_used=self.parser_name)
        blocks: list[BlockRecord] = []
        segments: list[NormalizedSegment] = []
        sheet_hints: list[dict[str, object]] = []
        segment_index = 0

        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            raw_rows = list(sheet.iter_rows(values_only=True))
            prepared_rows = [
                {"row_index": row_index, "values": _normalize_row(row)}
                for row_index, row in enumerate(raw_rows, start=1)
            ]
            regions = _sheet_regions(prepared_rows)
            if not regions:
                continue

            sheet_section_path = [document.title, sheet.title]
            sheet_heading_id = f"{document.document_id}-block-{len(blocks)}"
            blocks.append(
                BlockRecord(
                    block_id=sheet_heading_id,
                    block_type="heading",
                    text=sheet.title,
                    order_index=len(blocks),
                    title=sheet.title,
                    section_path=sheet_section_path,
                    citation_anchor=CitationAnchor(sheet_name=sheet.title, source_label=sheet.title),
                    heading_level=1,
                    metadata={"sheet_name": sheet.title, "sheet_index": sheet_index, "layout_hint": "sheet_heading"},
                )
            )

            region_hints: list[dict[str, object]] = []
            for region_index, region in enumerate(regions, start=1):
                region_rows = list(region)
                region_type = _region_type(region_rows)
                region_title = _region_title(sheet.title, region_rows, region_type, region_index)
                section_path = [*sheet_section_path, region_title]
                region_heading_id = f"{document.document_id}-block-{len(blocks)}"
                blocks.append(
                    BlockRecord(
                        block_id=region_heading_id,
                        block_type="heading",
                        text=region_title,
                        order_index=len(blocks),
                        title=region_title,
                        section_path=section_path,
                        citation_anchor=CitationAnchor(
                            sheet_name=sheet.title,
                            line_start=region_rows[0]["row_index"],
                            source_label=f"{sheet.title}:{region_rows[0]['row_index']}",
                        ),
                        heading_level=2,
                        metadata={
                            "sheet_name": sheet.title,
                            "sheet_index": sheet_index,
                            "region_index": region_index,
                            "region_type": region_type,
                            "layout_hint": "region_heading",
                        },
                    )
                )

                headers = _region_headers(region_rows, region_type)
                summary_text = _region_summary_text(region_rows, headers, region_type)
                summary_block_type = "table" if region_type == "table" else "paragraph"
                summary_block_id = f"{document.document_id}-block-{len(blocks)}"
                summary_anchor = CitationAnchor(
                    sheet_name=sheet.title,
                    line_start=region_rows[0]["row_index"],
                    line_end=region_rows[-1]["row_index"],
                    source_label=f"{sheet.title}:{region_rows[0]['row_index']}-{region_rows[-1]['row_index']}",
                )
                blocks.append(
                    BlockRecord(
                        block_id=summary_block_id,
                        block_type=summary_block_type,
                        text=summary_text,
                        order_index=len(blocks),
                        title=region_title,
                        section_path=section_path,
                        citation_anchor=summary_anchor,
                        metadata={
                            "sheet_name": sheet.title,
                            "sheet_index": sheet_index,
                            "region_index": region_index,
                            "region_type": region_type,
                            "headers": headers,
                            "row_start": region_rows[0]["row_index"],
                            "row_end": region_rows[-1]["row_index"],
                            "layout_hint": "table_summary" if region_type == "table" else "key_value_summary",
                        },
                    )
                )
                segments.append(
                    NormalizedSegment(
                        segment_id=f"{document.document_id}-seg-{segment_index}",
                        segment_type="sheet_summary",
                        text=summary_text,
                        title=region_title,
                        section_path=section_path,
                        citation_anchor=summary_anchor,
                        metadata={
                            "sheet_name": sheet.title,
                            "sheet_index": sheet_index,
                            "region_index": region_index,
                            "region_type": region_type,
                            "headers": headers,
                        },
                    )
                )
                segment_index += 1

                row_block_ids: list[str] = []
                for row in _region_data_rows(region_rows, region_type):
                    row_text = _region_row_text(row, headers, region_type)
                    if not row_text:
                        continue
                    row_index = int(row["row_index"])
                    block_id = f"{document.document_id}-block-{len(blocks)}"
                    row_block_ids.append(block_id)
                    anchor = CitationAnchor(sheet_name=sheet.title, line_start=row_index, source_label=f"{sheet.title}:{row_index}")
                    blocks.append(
                        BlockRecord(
                            block_id=block_id,
                            block_type="table_row",
                            text=row_text,
                            order_index=len(blocks),
                            title=region_title,
                            section_path=section_path,
                            citation_anchor=anchor,
                            metadata={
                                "sheet_name": sheet.title,
                                "sheet_index": sheet_index,
                                "region_index": region_index,
                                "region_type": region_type,
                                "row_index": row_index,
                                "headers": headers,
                                "layout_hint": "table_row" if region_type == "table" else "key_value_row",
                            },
                        )
                    )
                    segments.append(
                        NormalizedSegment(
                            segment_id=f"{document.document_id}-seg-{segment_index}",
                            segment_type="table_row",
                            text=row_text,
                            title=region_title,
                            section_path=section_path,
                            citation_anchor=anchor,
                            metadata={
                                "sheet_name": sheet.title,
                                "sheet_index": sheet_index,
                                "region_index": region_index,
                                "region_type": region_type,
                                "row_index": row_index,
                                "headers": headers,
                            },
                        )
                    )
                    segment_index += 1

                region_hints.append(
                    {
                        "region_index": region_index,
                        "region_type": region_type,
                        "region_title": region_title,
                        "section_path": section_path,
                        "heading_block_id": region_heading_id,
                        "summary_block_id": summary_block_id,
                        "row_block_ids": row_block_ids,
                        "headers": headers,
                        "row_start": region_rows[0]["row_index"],
                        "row_end": region_rows[-1]["row_index"],
                    }
                )

            sheet_hints.append(
                {
                    "sheet_name": sheet.title,
                    "sheet_index": sheet_index,
                    "section_path": sheet_section_path,
                    "heading_block_id": sheet_heading_id,
                    "regions": region_hints,
                }
            )

        document.blocks = blocks
        document.segments = segments
        document.metadata = {
            **document.metadata,
            "document_structure": "workbook_sheets",
            "chunking_hints": {
                "preferred_strategy": "xlsx_sheet_window",
                "structure": "workbook_sheets",
                "sheet_count": len(sheet_hints),
                "sheets": sheet_hints,
            },
        }
        document.language_tags = detect_languages(block.text for block in blocks if block.text)
        for block in document.blocks:
            block.language_tags = document.language_tags
        for segment in document.segments:
            segment.language_tags = document.language_tags
        return document


def _normalize_row(row) -> list[str]:
    return [normalize_text("" if value is None else str(value)) for value in row]


def _sheet_regions(rows: list[dict[str, object]]) -> list[list[dict[str, object]]]:
    regions: list[list[dict[str, object]]] = []
    current: list[dict[str, object]] = []
    for row in rows:
        values = row["values"]
        if _is_empty_row(values):
            if current:
                regions.append(current)
                current = []
            continue
        current.append(row)
    if current:
        regions.append(current)
    return regions


def _is_empty_row(values: list[str]) -> bool:
    return not any(value.strip() for value in values)


def _non_empty_values(values: list[str]) -> list[str]:
    return [value for value in values if value.strip()]


def _looks_numeric(value: str) -> bool:
    try:
        float(value.replace(",", ""))
    except ValueError:
        return False
    return True


def _region_type(rows: list[dict[str, object]]) -> str:
    if not rows:
        return "table"
    header_count = len(_non_empty_values(rows[0]["values"]))
    data_values = [_non_empty_values(row["values"]) for row in rows[1:]]
    data_counts = [len(values) for values in data_values]
    if header_count == 2 and data_values and all(count <= 2 for count in data_counts):
        numeric_value_column = any(any(_looks_numeric(value) for value in values[1:]) for values in data_values)
        if not numeric_value_column:
            return "key_value"
    if header_count >= 2 and data_counts and sum(1 for count in data_counts if count == header_count) >= max(1, len(data_counts) // 2):
        return "table"
    populated_counts = [len(_non_empty_values(row["values"])) for row in rows]
    mostly_two_column = sum(1 for count in populated_counts if 1 <= count <= 2) >= max(1, len(rows) - 1)
    first_values = [(_non_empty_values(row["values"]) or [""])[0] for row in rows]
    descriptive_first_values = sum(1 for value in first_values if value and not value.isdigit())
    if mostly_two_column and descriptive_first_values >= max(1, len(rows) - 1):
        return "key_value"
    return "table"


def _region_title(sheet_name: str, rows: list[dict[str, object]], region_type: str, region_index: int) -> str:
    if region_type == "key_value":
        first = _non_empty_values(rows[0]["values"])
        if first and len(first[0].split()) <= 6:
            return first[0]
        return f"{sheet_name} Details {region_index}"
    header_values = _non_empty_values(rows[0]["values"])
    if header_values:
        preview = ", ".join(header_values[:3])
        if len(preview) <= 48:
            return f"Table {region_index}: {preview}"
    return f"Table {region_index}"


def _region_headers(rows: list[dict[str, object]], region_type: str) -> list[str]:
    if not rows:
        return []
    first_values = list(rows[0]["values"])
    if region_type == "table":
        headers = []
        for index, value in enumerate(first_values, start=1):
            headers.append(value or f"column_{index}")
        return headers
    headers = []
    for row in rows:
        values = _non_empty_values(row["values"])
        if not values:
            continue
        headers.append(values[0])
    return headers[:12]


def _region_summary_text(rows: list[dict[str, object]], headers: list[str], region_type: str) -> str:
    if region_type == "key_value":
        lines = []
        for row in rows[:8]:
            values = _non_empty_values(row["values"])
            if not values:
                continue
            if len(values) == 1:
                lines.append(values[0])
            else:
                lines.append(f"{values[0]}: {values[1]}")
        return normalize_text("\n".join(lines))

    sample_lines = []
    for row in rows[1: min(len(rows), 6)]:
        row_line = _region_row_text(row, headers, "table")
        if row_line:
            sample_lines.append(row_line)
    summary_parts = []
    if headers:
        summary_parts.append(f"Headers: {', '.join(header for header in headers if header)}")
    if sample_lines:
        summary_parts.append("Sample rows:\n" + "\n".join(sample_lines))
    return normalize_text("\n\n".join(summary_parts)) or (f"Headers: {', '.join(headers)}" if headers else "Table")


def _region_data_rows(rows: list[dict[str, object]], region_type: str) -> list[dict[str, object]]:
    if region_type == "table" and len(rows) > 1:
        return rows[1:]
    return rows


def _region_row_text(row: dict[str, object], headers: list[str], region_type: str) -> str:
    values = list(row["values"])
    if region_type == "key_value":
        populated = _non_empty_values(values)
        if not populated:
            return ""
        if len(populated) == 1:
            return populated[0]
        return normalize_text(f"{populated[0]}: {populated[1]}")

    parts = []
    for header, value in zip(headers, values, strict=False):
        if not value:
            continue
        parts.append(f"{header}: {value}")
    return normalize_text(" | ".join(parts))
